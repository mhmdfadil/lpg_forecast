"""
Service untuk membangun file Excel hasil forecasting (multi-sheet, detail),
mengikuti struktur laporan pada notebook penelitian (Bab 11-13 ipynb acuan):
    1. Ringkasan Run
    2. Data Historis Harian (Train + Test)
    3. Uji Stasioneritas (ADF)
    4. ACF & PACF
    5. Parameter & Info Model ARIMA
    6. Parameter & Info Model SARIMA
    7. Prediksi ARIMA vs Aktual (Test)
    8. Prediksi SARIMA vs Aktual (Test)
    9. Peramalan Masa Depan vs Aktual 2026
    10. Evaluasi & Perbandingan Model
    11. Kesimpulan Model Terbaik
"""
import io
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="2A6EAF")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
TITLE_FONT = Font(bold=True, size=14, color="2A6EAF", name="Calibri")
SUBTITLE_FONT = Font(italic=True, size=10, color="666666", name="Calibri")
BEST_FILL = PatternFill("solid", fgColor="FFF3CD")
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)


def _style_header_row(ws, row=1):
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER


def _autosize(ws, max_width=45):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(length + 3, max_width)


def _write_df(ws, df: pd.DataFrame, start_row=1):
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=j, value=col)
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)
    _style_header_row(ws, start_row)
    return start_row + len(df) + 1


def build_forecast_excel(run, dataset_points, acf_before, acf_after,
                          pred_test_arima, pred_test_sarima,
                          pred_future_arima, pred_future_sarima) -> io.BytesIO:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ---------- Sheet 1: Ringkasan Run ----------
    ws = wb.create_sheet("1_Ringkasan")
    ws["A1"] = "LAPORAN HASIL FORECASTING PERMINTAAN LPG"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Metode: ARIMA & SARIMA — PT Pertamina Patra Niaga"
    ws["A2"].font = SUBTITLE_FONT
    info = [
        ("Nama Run", run.get("nama_run")),
        ("Wilayah / Scope", run.get("wilayah_scope")),
        ("Periode Data", f"{run.get('tanggal_mulai_data')} s.d. {run.get('tanggal_akhir_data')}"),
        ("Proporsi Train/Test", f"{float(run.get('train_test_split') or 0) * 100:.0f}% / {(1 - float(run.get('train_test_split') or 0)) * 100:.0f}%"),
        ("Periode Prediksi", f"{run.get('tanggal_mulai_prediksi')} s.d. {run.get('tanggal_akhir_prediksi')}"),
        ("Horizon (hari)", run.get("horizon_hari")),
        ("Dibandingkan Data Aktual 2026", "Ya" if run.get("bandingkan_aktual") else "Tidak"),
        ("Tanggal Proses", run.get("created_at")),
        ("", ""),
        ("MODEL TERBAIK", run.get("model_terbaik")),
        ("Skor ARIMA", f"{run.get('skor_arima')}/3"),
        ("Skor SARIMA", f"{run.get('skor_sarima')}/3"),
    ]
    r = 4
    for label, val in info:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)
        r += 1
    _autosize(ws)

    # ---------- Sheet 2: Data Harian (Train/Test) ----------
    ws = wb.create_sheet("2_Data_Harian")
    df_points = pd.DataFrame(dataset_points)
    if not df_points.empty:
        df_points = df_points[["tanggal", "nilai_aktual", "kelompok"]].rename(columns={
            "tanggal": "Tanggal", "nilai_aktual": "Total Penyaluran (Kg)", "kelompok": "Kelompok Data"
        })
    else:
        df_points = pd.DataFrame(columns=["Tanggal", "Total Penyaluran (Kg)", "Kelompok Data"])
    _write_df(ws, df_points)
    _autosize(ws)

    # ---------- Sheet 3: Uji Stasioneritas (ADF) ----------
    ws = wb.create_sheet("3_Uji_Stasioneritas")
    df_adf = pd.DataFrame([
        {
            "Tahap": "Sebelum Differencing",
            "ADF Statistic": run.get("adf_statistic_awal"),
            "P-Value": run.get("adf_pvalue_awal"),
            "Stasioner (p < 0.05)": "Ya" if run.get("adf_stasioner_awal") else "Tidak",
        },
        {
            "Tahap": f"Sesudah Differencing (d={run.get('differencing_order')})",
            "ADF Statistic": run.get("adf_statistic_diff"),
            "P-Value": run.get("adf_pvalue_diff"),
            "Stasioner (p < 0.05)": "Ya" if run.get("adf_stasioner_diff") else "Tidak",
        },
    ])
    _write_df(ws, df_adf)
    _autosize(ws)

    # ---------- Sheet 4: ACF & PACF ----------
    ws = wb.create_sheet("4_ACF_PACF")
    df_acf_b = pd.DataFrame(acf_before)
    df_acf_a = pd.DataFrame(acf_after)
    next_row = 1
    ws.cell(row=next_row, column=1, value="Sebelum Differencing").font = Font(bold=True)
    next_row += 1
    if not df_acf_b.empty:
        df_acf_b = df_acf_b[["lag", "acf_value", "pacf_value"]].rename(columns={
            "lag": "Lag", "acf_value": "ACF", "pacf_value": "PACF"
        })
        next_row = _write_df(ws, df_acf_b, next_row)
    next_row += 1
    if not df_acf_a.empty:
        ws.cell(row=next_row, column=1, value="Sesudah Differencing").font = Font(bold=True)
        next_row += 1
        df_acf_a = df_acf_a[["lag", "acf_value", "pacf_value"]].rename(columns={
            "lag": "Lag", "acf_value": "ACF", "pacf_value": "PACF"
        })
        _write_df(ws, df_acf_a, next_row)
    _autosize(ws)

    # ---------- Sheet 5: Parameter ARIMA ----------
    ws = wb.create_sheet("5_Parameter_ARIMA")
    df_arima = pd.DataFrame([
        {"Parameter": "Order (p,d,q)", "Nilai": f"({run.get('arima_p')},{run.get('arima_d')},{run.get('arima_q')})"},
        {"Parameter": "AIC", "Nilai": run.get("arima_aic")},
        {"Parameter": "BIC", "Nilai": run.get("arima_bic")},
        {"Parameter": "MAE (Data Test)", "Nilai": run.get("arima_mae_test")},
        {"Parameter": "RMSE (Data Test)", "Nilai": run.get("arima_rmse_test")},
        {"Parameter": "MAPE % (Data Test)", "Nilai": run.get("arima_mape_test")},
        {"Parameter": "MAE (vs Aktual 2026)", "Nilai": run.get("arima_mae_aktual")},
        {"Parameter": "RMSE (vs Aktual 2026)", "Nilai": run.get("arima_rmse_aktual")},
        {"Parameter": "MAPE % (vs Aktual 2026)", "Nilai": run.get("arima_mape_aktual")},
    ])
    _write_df(ws, df_arima)
    _autosize(ws)

    # ---------- Sheet 6: Parameter SARIMA ----------
    ws = wb.create_sheet("6_Parameter_SARIMA")
    df_sarima = pd.DataFrame([
        {"Parameter": "Order (p,d,q)", "Nilai": f"({run.get('sarima_p')},{run.get('sarima_d')},{run.get('sarima_q')})"},
        {"Parameter": "Seasonal Order (P,D,Q,s)", "Nilai": f"({run.get('sarima_pp')},{run.get('sarima_dd')},{run.get('sarima_qq')},{run.get('sarima_s')})"},
        {"Parameter": "AIC", "Nilai": run.get("sarima_aic")},
        {"Parameter": "BIC", "Nilai": run.get("sarima_bic")},
        {"Parameter": "MAE (Data Test)", "Nilai": run.get("sarima_mae_test")},
        {"Parameter": "RMSE (Data Test)", "Nilai": run.get("sarima_rmse_test")},
        {"Parameter": "MAPE % (Data Test)", "Nilai": run.get("sarima_mape_test")},
        {"Parameter": "MAE (vs Aktual 2026)", "Nilai": run.get("sarima_mae_aktual")},
        {"Parameter": "RMSE (vs Aktual 2026)", "Nilai": run.get("sarima_rmse_aktual")},
        {"Parameter": "MAPE % (vs Aktual 2026)", "Nilai": run.get("sarima_mape_aktual")},
    ])
    _write_df(ws, df_sarima)
    _autosize(ws)

    # ---------- Sheet 7: Prediksi ARIMA vs Aktual (Test) ----------
    ws = wb.create_sheet("7_Prediksi_ARIMA_Test")
    df_pa = pd.DataFrame(pred_test_arima)
    if not df_pa.empty:
        df_pa = df_pa[["tanggal", "nilai_aktual", "nilai_prediksi", "residual", "batas_bawah", "batas_atas"]].rename(columns={
            "tanggal": "Tanggal", "nilai_aktual": "Aktual", "nilai_prediksi": "Prediksi ARIMA",
            "residual": "Residual", "batas_bawah": "Batas Bawah (95%)", "batas_atas": "Batas Atas (95%)"
        })
    else:
        df_pa = pd.DataFrame(columns=["Tanggal", "Aktual", "Prediksi ARIMA", "Residual", "Batas Bawah (95%)", "Batas Atas (95%)"])
    _write_df(ws, df_pa)
    _autosize(ws)

    # ---------- Sheet 8: Prediksi SARIMA vs Aktual (Test) ----------
    ws = wb.create_sheet("8_Prediksi_SARIMA_Test")
    df_ps = pd.DataFrame(pred_test_sarima)
    if not df_ps.empty:
        df_ps = df_ps[["tanggal", "nilai_aktual", "nilai_prediksi", "residual", "batas_bawah", "batas_atas"]].rename(columns={
            "tanggal": "Tanggal", "nilai_aktual": "Aktual", "nilai_prediksi": "Prediksi SARIMA",
            "residual": "Residual", "batas_bawah": "Batas Bawah (95%)", "batas_atas": "Batas Atas (95%)"
        })
    else:
        df_ps = pd.DataFrame(columns=["Tanggal", "Aktual", "Prediksi SARIMA", "Residual", "Batas Bawah (95%)", "Batas Atas (95%)"])
    _write_df(ws, df_ps)
    _autosize(ws)

    # ---------- Sheet 9: Peramalan Masa Depan vs Aktual 2026 ----------
    ws = wb.create_sheet("9_Peramalan_vs_Aktual2026")
    df_fa = pd.DataFrame(pred_future_arima)
    df_fs = pd.DataFrame(pred_future_sarima)
    combined = None
    if not df_fa.empty:
        combined = df_fa[["tanggal", "nilai_prediksi", "batas_bawah", "batas_atas", "nilai_aktual_2026"]].rename(columns={
            "tanggal": "Tanggal", "nilai_prediksi": "Prediksi ARIMA",
            "batas_bawah": "ARIMA Batas Bawah", "batas_atas": "ARIMA Batas Atas",
            "nilai_aktual_2026": "Aktual 2026",
        })
    if not df_fs.empty:
        df_fs_ren = df_fs[["tanggal", "nilai_prediksi", "batas_bawah", "batas_atas"]].rename(columns={
            "tanggal": "Tanggal", "nilai_prediksi": "Prediksi SARIMA",
            "batas_bawah": "SARIMA Batas Bawah", "batas_atas": "SARIMA Batas Atas",
        })
        if combined is not None:
            combined = combined.merge(df_fs_ren, on="Tanggal", how="outer")
        else:
            combined = df_fs_ren
    if combined is None:
        combined = pd.DataFrame(columns=["Tanggal", "Prediksi ARIMA", "Prediksi SARIMA", "Aktual 2026"])
    else:
        cols_order = ["Tanggal", "Prediksi ARIMA", "ARIMA Batas Bawah", "ARIMA Batas Atas",
                      "Prediksi SARIMA", "SARIMA Batas Bawah", "SARIMA Batas Atas", "Aktual 2026"]
        combined = combined[[c for c in cols_order if c in combined.columns]]
    _write_df(ws, combined)
    _autosize(ws)

    # ---------- Sheet 10: Evaluasi & Perbandingan ----------
    ws = wb.create_sheet("10_Evaluasi_Perbandingan")
    df_eval = pd.DataFrame([
        {"Metrik": "MAE (Data Test)", "ARIMA": run.get("arima_mae_test"), "SARIMA": run.get("sarima_mae_test")},
        {"Metrik": "RMSE (Data Test)", "ARIMA": run.get("arima_rmse_test"), "SARIMA": run.get("sarima_rmse_test")},
        {"Metrik": "MAPE % (Data Test)", "ARIMA": run.get("arima_mape_test"), "SARIMA": run.get("sarima_mape_test")},
        {"Metrik": "MAE (vs Aktual 2026)", "ARIMA": run.get("arima_mae_aktual"), "SARIMA": run.get("sarima_mae_aktual")},
        {"Metrik": "RMSE (vs Aktual 2026)", "ARIMA": run.get("arima_rmse_aktual"), "SARIMA": run.get("sarima_rmse_aktual")},
        {"Metrik": "MAPE % (vs Aktual 2026)", "ARIMA": run.get("arima_mape_aktual"), "SARIMA": run.get("sarima_mape_aktual")},
        {"Metrik": "AIC", "ARIMA": run.get("arima_aic"), "SARIMA": run.get("sarima_aic")},
        {"Metrik": "BIC", "ARIMA": run.get("arima_bic"), "SARIMA": run.get("sarima_bic")},
    ])
    next_row = _write_df(ws, df_eval)
    ws.cell(row=next_row + 1, column=1, value="Skor Perbandingan").font = Font(bold=True)
    ws.cell(row=next_row + 2, column=1, value="ARIMA")
    ws.cell(row=next_row + 2, column=2, value=f"{run.get('skor_arima')}/3")
    ws.cell(row=next_row + 3, column=1, value="SARIMA")
    ws.cell(row=next_row + 3, column=2, value=f"{run.get('skor_sarima')}/3")
    _autosize(ws)

    # ---------- Sheet 11: Kesimpulan ----------
    ws = wb.create_sheet("11_Kesimpulan")
    ws["A1"] = "KESIMPULAN HASIL FORECASTING"
    ws["A1"].font = TITLE_FONT
    df_concl = pd.DataFrame([
        {"Item": "Model Terbaik", "Nilai": run.get("model_terbaik")},
        {"Item": "Order ARIMA", "Nilai": f"({run.get('arima_p')},{run.get('arima_d')},{run.get('arima_q')})"},
        {"Item": "Order SARIMA", "Nilai": f"({run.get('sarima_p')},{run.get('sarima_d')},{run.get('sarima_q')})({run.get('sarima_pp')},{run.get('sarima_dd')},{run.get('sarima_qq')})[{run.get('sarima_s')}]"},
        {"Item": "RMSE ARIMA (Test)", "Nilai": run.get("arima_rmse_test")},
        {"Item": "RMSE SARIMA (Test)", "Nilai": run.get("sarima_rmse_test")},
        {"Item": "Skor ARIMA", "Nilai": f"{run.get('skor_arima')}/3"},
        {"Item": "Skor SARIMA", "Nilai": f"{run.get('skor_sarima')}/3"},
        {"Item": "Wilayah / Scope", "Nilai": run.get("wilayah_scope")},
        {"Item": "Periode Data Historis", "Nilai": f"{run.get('tanggal_mulai_data')} s.d. {run.get('tanggal_akhir_data')}"},
        {"Item": "Periode Peramalan", "Nilai": f"{run.get('tanggal_mulai_prediksi')} s.d. {run.get('tanggal_akhir_prediksi')}"},
    ])
    _write_df(ws, df_concl, start_row=3)
    _autosize(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf